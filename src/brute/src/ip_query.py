"""IP 地理位置查询与 Excel 输出.

生成 ipcx.py 脚本并执行，将 xui.txt 结果转换为 xui.xlsx 带地理信息.
"""

import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def extract_host_port(line: str) -> str:
    """从行中提取 host:port."""
    match = re.search(r"https?://([^/\s]+)", line)
    if match:
        return match.group(1)
    return line.strip()


def get_ip_info(ip_port: str, retries: int = 3) -> list:
    """调用 ip-api.com 获取 IP 地理信息."""
    if ":" in ip_port:
        ip, port = ip_port.split(":", 1)
    else:
        ip = ip_port.strip()
        port = ""

    url = f"http://ip-api.com/json/{ip}?fields=country,regionName,city,isp"
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                info = resp.json()
                return [
                    f"{ip}:{port}" if port else ip,
                    info.get("country", "N/A"),
                    info.get("regionName", "N/A"),
                    info.get("city", "N/A"),
                    info.get("isp", "N/A"),
                ]
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(1)
    return [f"{ip}:{port}" if port else ip, "N/A", "N/A", "N/A", "N/A"]


def format_time(seconds: float) -> str:
    minutes = int(seconds) // 60
    secs = int(seconds) % 60
    return f"{minutes}分钟{secs}秒"


def adjust_column_width(ws):
    """自动调整 Excel 列宽."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2


def extract_ip_port(url: str) -> str:
    """从 URL 或 ip:port 文本中提取主机部分."""
    match = re.search(r"https?://([^/\s]+)", url)
    if match:
        return match.group(1)
    if ":" in url:
        return url.split()[0]
    return url.split()[0]


def process_ip_port_file(input_file: str = "xui.txt", output_excel: str = "xui.xlsx") -> str:
    """主处理函数：读取 xui.txt，写入 xui.xlsx."""
    input_path = Path(input_file)
    if not input_path.exists():
        print(f"⚠️ 输入文件 {input_file} 不存在，跳过 Excel 生成")
        return ""

    lines = [line.strip() for line in input_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    total_tasks = len(lines)
    completed_tasks = 0
    start_time = time.time()

    headers = ["原始地址", "IP/域名:端口", "用户名", "密码", "国家", "地区", "城市", "ISP"]

    output_path = Path(output_excel)
    if output_path.exists():
        output_path.unlink()

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(str(output_path))

    for line in lines:
        parts = line.split()
        if len(parts) >= 3:
            addr, user, passwd = parts[:3]
        else:
            addr = parts[0]
            user = passwd = ""

        ip_port = extract_ip_port(addr)
        ip_info = get_ip_info(ip_port)
        row = [addr, ip_port, user, passwd] + ip_info[1:]

        wb = load_workbook(str(output_path))
        ws = wb.active
        ws.append(row)
        adjust_column_width(ws)
        wb.save(str(output_path))

        completed_tasks += 1
        elapsed = time.time() - start_time
        avg_time = elapsed / completed_tasks
        remaining = avg_time * (total_tasks - completed_tasks)
        percent = (completed_tasks / total_tasks) * 100
        eta = format_time(remaining)
        print(
            f"\r处理进度: {completed_tasks}/{total_tasks} ({percent:.2f}%) "
            f"预计剩余时间: {eta}",
            end="", flush=True,
        )
        time.sleep(1.5)

    print("\n全部处理完成！")
    return str(output_path)


def generate_ipcx_script(output_path: str = "ipcx.py"):
    """生成独立的 ipcx.py 脚本文件."""
    script = '''"""IP 地理位置查询与 Excel 生成脚本 —— 由 x-ui 生成器自动生成."""
import time
import re
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def extract_host_port(line):
    match = re.search(r"https?://([^/\\s]+)", line)
    if match:
        return match.group(1)
    return line.strip()


def get_ip_info(ip_port, retries=3):
    if ":" in ip_port:
        ip, port = ip_port.split(":", 1)
    else:
        ip = ip_port.strip()
        port = ""
    url = f"http://ip-api.com/json/{ip}?fields=country,regionName,city,isp"
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                info = resp.json()
                return [
                    f"{ip}:{port}" if port else ip,
                    info.get("country", "N/A"),
                    info.get("regionName", "N/A"),
                    info.get("city", "N/A"),
                    info.get("isp", "N/A"),
                ]
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(1)
    return [f"{ip}:{port}" if port else ip, "N/A", "N/A", "N/A", "N/A"]


def format_time(seconds):
    minutes = int(seconds) // 60
    secs = int(seconds) % 60
    return f"{minutes}分钟{secs}秒"


def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2


def extract_ip_port(url):
    match = re.search(r"https?://([^/\\s]+)", url)
    if match:
        return match.group(1)
    if ":" in url:
        return url.split()[0]
    return url.split()[0]


def process(input_file="xui.txt", output_excel="xui.xlsx"):
    import os
    with open(input_file, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]
    total = len(lines)
    completed = 0
    start = time.time()

    headers = ["原始地址", "IP/域名:端口", "用户名", "密码", "国家", "地区", "城市", "ISP"]

    if os.path.exists(output_excel):
        os.remove(output_excel)

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(output_excel)

    for line in lines:
        parts = line.split()
        addr = parts[0]
        user = parts[1] if len(parts) >= 2 else ""
        passwd = parts[2] if len(parts) >= 3 else ""

        ip_port = extract_ip_port(addr)
        ip_info = get_ip_info(ip_port)
        row = [addr, ip_port, user, passwd] + ip_info[1:]

        wb = load_workbook(output_excel)
        ws = wb.active
        ws.append(row)
        adjust_column_width(ws)
        wb.save(output_excel)

        completed += 1
        elapsed = time.time() - start
        avg = elapsed / completed
        remaining = avg * (total - completed)
        percent = (completed / total) * 100
        eta = format_time(remaining)
        print(f"\\r处理进度: {completed}/{total} ({percent:.2f}%) 预计剩余: {eta}", end="", flush=True)
        time.sleep(1.5)
    print("\\n全部处理完成！")


if __name__ == "__main__":
    process()
'''
    Path(output_path).write_text(script, encoding="utf-8")
    return output_path
